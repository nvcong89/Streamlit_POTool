"""CSV schemas, loaders, and column mapping."""

from __future__ import annotations

from io import StringIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd

from src.analysis.section_analysis import MC_RESULT_COLUMNS

SECTION_INPUT_FIELDS = [
    "D",
    "D_bearing",
    "phi_main",
    "n_bar",
    "phi_link",
    "cover",
    "f_ce",
    "e_c0",
    "e_cu0",
    "e_spall",
    "r_c",
    "f_cu",
    "e_cc",
    "e_ccu",
    "r_cc",
    "f_cc",
    "e_ye",
    "e_sh",
    "e_smd",
    "f_ye",
    "f_ue",
    "Es",
    "e_ye_bearing",
    "e_sh_bearing",
    "e_smd_bearing",
    "f_ye_bearing",
    "f_ue_bearing",
    "Es_bearing",
    "P_start",
    "P_end",
    "e_ccmax_model",
    "e_cmax_model",
    "e_smax_model",
    "code",
    "Pmin",
    "Pmax",
    "Lp",
    "ec_MD",
    "ec_CD",
    "ec_ED",
    "es_MD",
    "es_CD",
    "es_ED",
]

MC_REQUIRED_FIELDS = ["P", "M", "curvature"]
MC_OPTIONAL_FIELDS = [c for c in MC_RESULT_COLUMNS if c not in MC_REQUIRED_FIELDS]

MATERIAL_CURVE_FIELDS = ["material", "strain", "stress"]

DEFAULT_SECTION_VALUES: Dict[str, float | str] = {
    "D": 680.0,
    "D_bearing": 700.0,
    "phi_main": 32.0,
    "n_bar": 12.0,
    "phi_link": 10.0,
    "cover": 40.0,
    "f_ce": 40.0,
    "e_c0": 0.002,
    "e_cu0": 0.004,
    "e_spall": 0.006,
    "r_c": 4.0,
    "f_cu": 0.0,
    "e_cc": 0.0035,
    "e_ccu": 0.015,
    "r_cc": 1.5,
    "f_cc": 48.0,
    "e_ye": 0.00275,
    "e_sh": 0.005,
    "e_smd": 0.08,
    "f_ye": 500.0,
    "f_ue": 650.0,
    "Es": 200000.0,
    "e_ye_bearing": 0.002,
    "e_sh_bearing": 0.005,
    "e_smd_bearing": 0.05,
    "f_ye_bearing": 355.0,
    "f_ue_bearing": 510.0,
    "Es_bearing": 200000.0,
    "P_start": -3184.8,
    "P_end": 14182.8,
    "e_ccmax_model": 0.015,
    "e_cmax_model": 0.025,
    "e_smax_model": 0.12,
    "code": "ASCE 61-14",
    "Pmin": -10616.0,
    "Pmax": 28365.0,
    "Lp": 1.083,
    "ec_MD": 0.004973,
    "ec_CD": 0.008574,
    "ec_ED": 0.012862,
    "es_MD": 0.0145,
    "es_CD": 0.025,
    "es_ED": 0.05,
}


def load_csv_upload(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".xlsm"):
        return parse_mc_data_xlsm(uploaded_file)["mc_results"]
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return pd.read_excel(uploaded_file)
    return pd.read_csv(uploaded_file)


def _cell_float(ws, addr: str, default: float = 0.0) -> float:
    val = ws[addr].value
    if val is None:
        return default
    return float(val)


def _open_mc_data_workbook(source):
    """Open workbook and return MC_Data worksheet."""
    import openpyxl

    if hasattr(source, "read"):
        if hasattr(source, "seek"):
            source.seek(0)
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)

    if "MC_Data" not in wb.sheetnames:
        wb.close()
        raise ValueError("Workbook must contain an 'MC_Data' sheet.")
    return wb, wb["MC_Data"]


def _parse_mc_data_input_ws(ws) -> Dict:
    """
    Parse MC_Data input region (A1:S20, values rows 4–19).

    Cell map matches Excel MC Analysis workbook layout.
    """
    section = dict(DEFAULT_SECTION_VALUES)
    section.update(
        {
            "D": _cell_float(ws, "B4"),
            "D0": _cell_float(ws, "B5"),
            "phi_main": _cell_float(ws, "B6"),
            "n_bar": _cell_float(ws, "B7"),
            "phi_link": _cell_float(ws, "B8"),
            "cover": _cell_float(ws, "B9"),
            "D_bearing": _cell_float(ws, "B10"),
            "f_ce": _cell_float(ws, "E4"),
            "e_c0": _cell_float(ws, "H4"),
            "e_cu0": _cell_float(ws, "H5"),
            "e_spall": _cell_float(ws, "H6"),
            "r_c": _cell_float(ws, "H7"),
            "f_cu": _cell_float(ws, "H8"),
            "e_cc": _cell_float(ws, "K4"),
            "e_ccu": _cell_float(ws, "K5"),
            "r_cc": _cell_float(ws, "K6"),
            "f_cc": _cell_float(ws, "K7"),
            "e_ye": _cell_float(ws, "N4"),
            "e_sh": _cell_float(ws, "N5"),
            "e_smd": _cell_float(ws, "N6"),
            "f_ye": _cell_float(ws, "N7"),
            "f_ue": _cell_float(ws, "N8"),
            "Es": _cell_float(ws, "N9"),
            "e_ye_bearing": _cell_float(ws, "Q4"),
            "e_sh_bearing": _cell_float(ws, "Q5"),
            "e_smd_bearing": _cell_float(ws, "Q6"),
            "f_ye_bearing": _cell_float(ws, "Q7"),
            "f_ue_bearing": _cell_float(ws, "Q8"),
            "Es_bearing": _cell_float(ws, "Q9"),
            "Lp": _cell_float(ws, "H12"),
            "ec_MD": _cell_float(ws, "B13"),
            "ec_CD": _cell_float(ws, "D13"),
            "ec_ED": _cell_float(ws, "F13"),
            "es_MD": _cell_float(ws, "B14"),
            "es_CD": _cell_float(ws, "D14"),
            "es_ED": _cell_float(ws, "F14"),
            "Pmin": _cell_float(ws, "B16"),
            "Pmax": _cell_float(ws, "B17"),
            "P_start": _cell_float(ws, "B18"),
            "P_end": _cell_float(ws, "B19"),
            "e_ccmax_model": _cell_float(ws, "E16"),
            "e_cmax_model": _cell_float(ws, "E17"),
            "e_smax_model": _cell_float(ws, "E18"),
        }
    )
    return _coerce_section_types(section)


def parse_mc_data_input(source) -> Dict:
    """
    Parse section input only from MC_Data sheet (A1:S20).

    Returns section_input dict suitable for run_mc_analysis().
    """
    wb, ws = _open_mc_data_workbook(source)
    try:
        return _parse_mc_data_input_ws(ws)
    finally:
        wb.close()


def parse_mc_data_xlsm(source, inputs_only: bool = False) -> Dict:
    """
    Parse MC_Data sheet from Excel workbook (.xlsm/.xlsx).

    Returns dict with keys: section_input, mc_results, pm_summary (optional).

    When inputs_only=True, mc_results and pm_summary are None.
    """
    wb, ws = _open_mc_data_workbook(source)
    section = _parse_mc_data_input_ws(ws)

    if inputs_only:
        wb.close()
        return {
            "section_input": section,
            "mc_results": None,
            "pm_summary": None,
        }

    mc_rows = []
    for row in ws.iter_rows(min_row=25, min_col=1, max_col=9, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        if str(row[0]).lower() in ("no.", "output"):
            continue
        try:
            mc_rows.append(
                {
                    "Point_Name": str(row[0]),
                    "P": float(row[1] or 0),
                    "M": float(row[2] or 0),
                    "e_cc": float(row[3] or 0),
                    "e_c": float(row[4] or 0),
                    "e_s": float(row[5] or 0),
                    "x": float(row[6] or 0),
                    "curvature": float(row[7] or 0),
                    "F_error": float(row[8] or 0),
                }
            )
        except (TypeError, ValueError):
            continue

    pm_rows = []
    for row in ws.iter_rows(min_row=25, min_col=11, max_col=22, values_only=True):
        if row[0] is None or str(row[0]).strip() == "":
            continue
        if str(row[0]).lower() == "point":
            continue
        try:
            point_name = str(row[0])
            pm_rows.append(
                {
                    "Point": point_name,
                    "P": float(row[1] or 0),
                    "Mp": float(row[2] or 0),
                    "Fy": float(row[3] or 0),
                    "FmOLE": float(row[4] or 0),
                    "FmCLE": float(row[5] or 0),
                    "FmDE": float(row[6] or 0),
                    "Fultimate": float(row[7] or 0),
                    "qp_mOLE": float(row[8] or 0),
                    "qp_mCLE": float(row[9] or 0),
                    "qp_mDE": float(row[10] or 0),
                    "qUltimate": float(row[11] or 0),
                }
            )
            if len(pm_rows) >= 60:
                break
        except (TypeError, ValueError):
            continue

    wb.close()
    return {
        "section_input": section,
        "mc_results": pd.DataFrame(mc_rows, columns=MC_RESULT_COLUMNS),
        "pm_summary": pd.DataFrame(pm_rows) if pm_rows else None,
    }


def parse_section_input(df: pd.DataFrame, column_map: Optional[Dict[str, str]] = None) -> Dict:
    """Parse section input from key-value or wide single-row CSV."""
    if column_map:
        df = df.rename(columns=column_map)

    cols_lower = {c.lower().strip(): c for c in df.columns}

    # Key-value format: parameter, value
    if "parameter" in cols_lower and "value" in cols_lower:
        pcol = cols_lower["parameter"]
        vcol = cols_lower["value"]
        data = {}
        for _, row in df.iterrows():
            key = str(row[pcol]).strip()
            if key in SECTION_INPUT_FIELDS or key in DEFAULT_SECTION_VALUES:
                val = row[vcol]
                if key == "code":
                    data[key] = str(val)
                else:
                    data[key] = float(val)
        for k, v in DEFAULT_SECTION_VALUES.items():
            data.setdefault(k, v)
        return _coerce_section_types(data)

    # Wide single-row
    if len(df) == 1:
        row = df.iloc[0]
        data = {}
        for field in SECTION_INPUT_FIELDS:
            if field in df.columns:
                val = row[field]
                data[field] = str(val) if field == "code" else float(val)
            elif field.lower() in cols_lower:
                orig = cols_lower[field.lower()]
                val = row[orig]
                data[field] = str(val) if field == "code" else float(val)
        for k, v in DEFAULT_SECTION_VALUES.items():
            data.setdefault(k, v)
        return _coerce_section_types(data)

    raise ValueError(
        "Section input must be key-value (parameter, value) or a single wide row. "
        f"Got {len(df)} rows with columns: {list(df.columns)}"
    )


def _coerce_section_types(data: Dict) -> Dict:
    out = dict(data)
    out["n_bar"] = int(float(out.get("n_bar", 20)))
    for key in SECTION_INPUT_FIELDS:
        if key == "code":
            out.setdefault(key, "ASCE 61-14")
        elif key != "code" and key in out:
            out[key] = float(out[key])
        elif key != "code":
            out[key] = float(DEFAULT_SECTION_VALUES[key])  # type: ignore[arg-type]
    return out


def parse_mc_results(df: pd.DataFrame, column_map: Optional[Dict[str, str]] = None) -> pd.DataFrame:
    if column_map:
        df = df.rename(columns=column_map)
    missing = [f for f in MC_REQUIRED_FIELDS if f not in df.columns]
    if missing:
        raise ValueError(f"Missing required M-φ columns: {missing}")
    out = df.copy()
    for col in MC_RESULT_COLUMNS:
        if col not in out.columns:
            if col == "Point_Name":
                out[col] = [f"Pt-{i}" for i in range(len(out))]
            elif col in ("e_cc", "e_c", "e_s", "x", "F_error"):
                out[col] = 0.0
            else:
                continue
    return out[MC_RESULT_COLUMNS]


def parse_material_curves(df: pd.DataFrame, column_map: Optional[Dict[str, str]] = None) -> pd.DataFrame:
    if column_map:
        df = df.rename(columns=column_map)
    missing = [f for f in MATERIAL_CURVE_FIELDS if f not in df.columns]
    if missing:
        raise ValueError(f"Missing material curve columns: {missing}")
    return df[MATERIAL_CURVE_FIELDS].copy()


def build_column_mapper(
    uploaded_columns: List[str], expected_fields: List[str]
) -> Tuple[Dict[str, str], List[str]]:
    """Auto-map columns by case-insensitive name match; return map and unmapped required."""
    mapping: Dict[str, str] = {}
    lower_to_orig = {c.lower().strip(): c for c in uploaded_columns}
    for field in expected_fields:
        if field in uploaded_columns:
            mapping[field] = field
        elif field.lower() in lower_to_orig:
            mapping[lower_to_orig[field.lower()]] = field
    unmapped = [f for f in expected_fields if f not in mapping.values()]
    return mapping, unmapped


def template_section_kv_csv() -> str:
    lines = ["parameter,value"]
    for k, v in DEFAULT_SECTION_VALUES.items():
        lines.append(f"{k},{v}")
    return "\n".join(lines)


def template_mc_results_csv() -> str:
    return ",".join(MC_RESULT_COLUMNS) + "\nP0-1,-1000,100,0.001,0.002,0.001,600,0.003,0.5\n"


def read_template(name: str) -> str:
    path = Path(__file__).resolve().parents[2] / "templates" / name
    if path.exists():
        return path.read_text(encoding="utf-8")
    if name == "section_input.csv":
        return template_section_kv_csv()
    if name == "mc_results.csv":
        return template_mc_results_csv()
    raise FileNotFoundError(name)


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")
